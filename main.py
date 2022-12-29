from openpyxl import load_workbook

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
load_wb = load_workbook(r"C:\Users\gpfks\OneDrive\바탕 화면\taxbill\exp.xlsx", data_only=True)
# 시트 이름으로 불러오기 
load_ws = load_wb['Sheet']
# 셀 주소로 값 출력
print(load_ws.cell(row=7, column=3).value)
# 시트의 최대 행과 열의 값
max_ro = load_ws.max_row
max_col = load_ws.max_column
print("maxro : " , max_ro)
print("max_col : " , max_col)

# 값을 저장할 딕셔너리 명 : dic
# 순서 : 코드, 상호명, 사업자번호, 금액, 배수
dic = {}
for num in range(2,max_ro-1):
    dic[num] = [load_ws.cell(row=num, column=1).value, 
    load_ws.cell(row=num, column=3).value, 
    load_ws.cell(row=num, column=5).value.replace("-", ""),
    load_ws.cell(row=num, column=10).value, 1]
print("사업자 정보 가져오기 완료")

# form 파일 가져오기
write_wb = load_workbook(r"C:\Users\gpfks\OneDrive\바탕 화면\taxbill\form.xlsx", data_only=True)
# 시트 이름으로 불러오기 
write_ws = write_wb['엑셀업로드양식']
print(write_ws.cell(row=1, column=2))

# 사업자 정보 입력하기
print("시작")
for num in range(2,max_ro-1):
    write_ws.cell(5+num, 1, '05')
    write_ws.cell(5+num, 2, 20221231)
    write_ws.cell(5+num, 3, dic[num][2])
    write_ws.cell(5+num, 5, dic[num][1])
    # 공급받는자 성명 
    # write_ws.cell(7+num, garo+5, )
    write_ws.cell(5+num, 12, dic[num][3] * dic[num][4])
    write_ws.cell(5+num, 46, '02')

write_wb.save(r"C:\Users\gpfks\OneDrive\바탕 화면\taxbill\form.xlsx")